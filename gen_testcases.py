# -*- coding: utf-8 -*-
"""Sinh file Excel danh sách test case cho báo cáo kiểm thử TECHACADEMY.
Chạy: python gen_testcases.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Mỗi dòng: [Mã TC, Nhóm, Mục tiêu, Các bước thực hiện, Kết quả mong đợi, Loại, Kết quả]
data = [
    # ----- Feature 1: Tìm kiếm -----
    ["TC-01", "Tìm kiếm", "Tìm kiếm với từ khóa chính xác", "1. Mở trang chủ\n2. Nhập 'Cơ Bản' vào ô tìm kiếm", "Hiển thị đúng 1 khóa học 'Golang Cơ Bản'", "Bình thường", "PASS"],
    ["TC-02", "Tìm kiếm", "Không phân biệt chữ hoa/thường", "1. Nhập 'golang'\n2. Nhập 'GOLANG'", "Cả hai trả về 2 khóa học giống nhau", "Bình thường", "PASS"],
    ["TC-03", "Tìm kiếm", "Tìm với từ khóa không tồn tại", "1. Nhập 'Python Django Machine Learning'", "Hiển thị 'Không tìm thấy khóa học nào phù hợp'", "Bình thường", "PASS"],
    ["TC-04", "Tìm kiếm", "Ô tìm kiếm để trống", "1. Xóa trắng ô tìm kiếm", "Hiển thị đầy đủ 6 khóa học", "Bình thường", "PASS"],
    ["TC-05", "Tìm kiếm", "Tìm kiếm chỉ chứa khoảng trắng", "1. Nhập '   '", "Xử lý an toàn, hiển thị thông báo không tìm thấy", "Bình thường", "PASS"],
    ["TC-06", "Tìm kiếm", "Chống tấn công SQL Injection", "1. Nhập các payload SQL độc hại vào ô tìm kiếm", "Không crash, không lộ dữ liệu, DB nguyên vẹn (6 khóa)", "Bình thường", "PASS"],
    ["TC-07", "Tìm kiếm", "Lọc theo danh mục", "1. Click danh mục 'Go'", "Hiển thị đúng 2 khóa học thuộc Go", "Bình thường", "PASS"],
    ["TC-08", "Tìm kiếm", "Lọc theo mức học phí", "1. Click 'Miễn phí'\n2. Click 'Có phí'", "Miễn phí: 2 khóa; Có phí: 4 khóa", "Bình thường", "PASS"],
    ["TC-09", "Tìm kiếm", "Tìm với từ khóa cực dài (>200 ký tự)", "1. Nhập chuỗi 210 ký tự", "Đáng lẽ báo lỗi thân thiện; THỰC TẾ backend trả HTTP 500", "Lỗi cố ý", "FAIL"],
    ["TC-61", "Tìm kiếm", "Lọc kết hợp danh mục + học phí", "1. Click danh mục 'Go'\n2. Click 'Có phí'", "Đáng lẽ 1 khóa (Go có phí); THỰC TẾ trả 2 (bỏ qua lọc giá)", "Lỗi cố ý", "FAIL"],

    # ----- Feature 2: Đánh giá -----
    ["TC-10", "Đánh giá", "Gửi đánh giá thành công", "1. Đăng nhập\n2. Mở chi tiết khóa 1\n3. Chọn 4 sao, nhập bình luận, Gửi", "Ghi nhận thành công, bình luận xuất hiện trong danh sách", "Bình thường", "PASS"],
    ["TC-11", "Đánh giá", "Gửi đánh giá khi chưa đăng nhập", "1. Mở chi tiết khóa 1 (chưa login)", "Form bị vô hiệu hóa, hiện cảnh báo đăng nhập", "Bình thường", "PASS"],
    ["TC-12", "Đánh giá", "Đánh giá chỉ có sao, bình luận trống", "1. Chọn 5 sao, bỏ trống bình luận, Gửi", "Gửi thành công", "Bình thường", "PASS"],
    ["TC-13", "Đánh giá", "Biên: đánh giá với 0 sao", "1. Gọi API gửi review rating=0", "API trả 400 'Vui lòng chọn số sao từ 1 đến 5'", "Bình thường", "PASS"],
    ["TC-14", "Đánh giá", "Bình luận chỉ có khoảng trắng", "1. Chọn 4 sao, nhập '      ', Gửi", "Báo lỗi 'Bình luận phải có ít nhất 3 ký tự'", "Bình thường", "PASS"],
    ["TC-15", "Đánh giá", "Bình luận < 3 ký tự", "1. Nhập 'Hi', Gửi", "Báo lỗi tối thiểu 3 ký tự", "Bình thường", "PASS"],
    ["TC-16", "Đánh giá", "Bình luận > 500 ký tự", "1. Nhập 505 ký tự, Gửi", "Báo lỗi vượt quá 500 ký tự", "Bình thường", "PASS"],
    ["TC-17", "Đánh giá", "Đánh giá mới ở đầu danh sách", "1. Gửi đánh giá mới", "Bình luận mới hiển thị trên cùng", "Bình thường", "PASS"],
    ["TC-18", "Đánh giá", "SQL Injection qua bình luận", "1. Gửi payload SQL trong nội dung bình luận", "Lưu nguyên dạng text, không gây lỗi hệ thống", "Bình thường", "PASS"],
    ["TC-19", "Đánh giá", "Chặn đánh giá nhiều lần", "1. Login user đã đánh giá khóa 1\n2. Gửi tiếp đánh giá", "Báo lỗi 'Bạn đã đánh giá khóa học này rồi'", "Bình thường", "PASS"],
    ["TC-20", "Đánh giá", "Tính điểm trung bình", "1. Thêm review 1 sao cho khóa 1 (đang 4.5)", "Điểm trung bình cập nhật xuống 3.3", "Bình thường", "PASS"],
    ["TC-21", "Đánh giá", "Reset form khi đóng/mở lại modal", "1. Nhập dở review, đóng modal, mở lại", "Bình luận trống, số sao về mặc định 5", "Bình thường", "PASS"],
    ["TC-22", "Đánh giá", "Bình luận chứa emoji", "1. Gửi bình luận có emoji", "Lưu và hiển thị emoji chính xác", "Bình thường", "PASS"],
    ["TC-23", "Đánh giá", "Spam click nút Gửi đánh giá", "1. Click nút Gửi 2 lần liên tiếp", "Đáng lẽ 1 review; THỰC TẾ tạo 2 bản ghi trùng (race condition)", "Lỗi cố ý", "FAIL"],
    ["TC-24", "Đánh giá", "Lỗ hổng XSS qua bình luận", "1. Gửi payload <img onerror=...>", "Đáng lẽ phải sanitize; THỰC TẾ script thực thi (window.xssDetected=true)", "Lỗi cố ý", "FAIL"],
    ["TC-62", "Đánh giá", "Điểm trung bình không được làm tròn", "1. Thêm review 1 sao khóa 1 → TB (5+4+1)/3", "Đáng lẽ '3.3'; THỰC TẾ hiển thị '3.3333333333333335'", "Lỗi cố ý", "FAIL"],

    # ----- Feature 3: Đăng ký -----
    ["TC-25", "Đăng ký", "Đăng ký khóa miễn phí", "1. Đăng nhập\n2. Click Đăng ký khóa 1 (free)", "Thành công, badge 'Đã đăng ký'", "Bình thường", "PASS"],
    ["TC-26", "Đăng ký", "Đăng ký khóa có phí qua modal thanh toán", "1. Click Đăng ký khóa 2\n2. Xác nhận thanh toán", "Mở modal QR, đăng ký thành công, hiện badge", "Bình thường", "PASS"],
    ["TC-27", "Đăng ký", "Đăng ký khi chưa đăng nhập", "1. Click Đăng ký (chưa login)", "Hiện modal đăng nhập; login xong tiếp tục được", "Bình thường", "PASS"],
    ["TC-28", "Đăng ký", "Chặn đăng ký trùng", "1. Đăng ký khóa 1\n2. Gọi API đăng ký lại", "API trả 400 'Bạn đã đăng ký khóa học này rồi'", "Bình thường", "PASS"],
    ["TC-29", "Đăng ký", "Khóa đã đăng ký hiện ở Dashboard", "1. Đăng ký khóa 1\n2. Mở tab Khóa của tôi", "Khóa 1 xuất hiện trong danh sách", "Bình thường", "PASS"],
    ["TC-30", "Đăng ký", "Hủy đăng ký khóa học", "1. Đăng ký khóa 1\n2. Hủy đăng ký", "Khóa biến mất khỏi Dashboard", "Bình thường", "PASS"],
    ["TC-31", "Đăng ký", "Giới hạn số học viên (khóa 5 tối đa 2)", "1. 2 tài khoản đăng ký khóa 5\n2. Tài khoản thứ 3 đăng ký", "Báo 'Khóa học đã đầy học viên'", "Bình thường", "PASS"],
    ["TC-32", "Đăng ký", "Áp mã giảm giá 50% (GIAM50)", "1. Đăng ký khóa 2\n2. Áp GIAM50", "Giá còn $25.00", "Bình thường", "PASS"],
    ["TC-33", "Đăng ký", "Áp mã giảm giá 100% (FREE100)", "1. Áp FREE100", "Giá về $0.00", "Bình thường", "PASS"],
    ["TC-34", "Đăng ký", "Mã giảm giá đã hết hạn (EXPIRED)", "1. Áp mã EXPIRED", "Báo 'Mã giảm giá đã hết hạn'", "Bình thường", "PASS"],
    ["TC-35", "Đăng ký", "Mã giảm giá không hợp lệ", "1. Áp mã 'INVALIDCODE'", "Báo 'Mã giảm giá không hợp lệ'", "Bình thường", "PASS"],
    ["TC-36", "Đăng ký", "Đổi mã giảm giá động", "1. Áp FREE100 → $0.00\n2. Đổi GIAM50 → $25.00", "Giá cập nhật đúng theo mã mới nhất", "Bình thường", "PASS"],
    ["TC-37", "Đăng ký", "Đóng modal thanh toán bằng overlay", "1. Mở checkout khóa 2\n2. Click overlay đóng", "Trạng thái sạch, không đăng ký (0 khóa)", "Bình thường", "PASS"],
    ["TC-38", "Đăng ký", "Nút 'Vào học ngay' cho khóa đã sở hữu", "1. Đăng ký khóa 1\n2. Mở Dashboard", "Hiện nút 'Vào học ngay'", "Bình thường", "PASS"],
    ["TC-39", "Đăng ký", "'Vào học ngay' mở chi tiết khóa học", "1. Click 'Vào học ngay'", "Mở modal chi tiết và nội dung học", "Bình thường", "PASS"],
    ["TC-40", "Đăng ký", "Quay lại tab Khám phá", "1. Vào Dashboard\n2. Click Khám phá", "Hiển thị đầy đủ 6 khóa", "Bình thường", "PASS"],
    ["TC-41", "Đăng ký", "Đăng ký nhiều khóa học", "1. Đăng ký khóa 1 và khóa 3", "Dashboard có 2 khóa", "Bình thường", "PASS"],
    ["TC-42", "Đăng ký", "Hủy thanh toán không đăng ký", "1. Mở checkout khóa 2\n2. Click Hủy thanh toán", "Không đăng ký (Dashboard 0 khóa)", "Bình thường", "PASS"],
    ["TC-43", "Đăng ký", "Áp GIAM20 đồng bộ Frontend/Backend", "1. Áp GIAM20\n2. Thanh toán", "Giá $39.99, API trả 201", "Bình thường", "PASS"],
    ["TC-44", "Đăng ký", "FREE100 bỏ qua màn quét QR", "1. Áp FREE100\n2. Thanh toán", "Bỏ qua QR, đăng ký thành công", "Bình thường", "PASS"],
    ["TC-63", "Đăng ký", "Tiền giảm giá không được làm tròn", "1. Áp GIAM20 khóa 2\n2. Thanh toán", "Đáng lẽ $39.99; THỰC TẾ '$39.992000000000004' trong thông báo", "Lỗi cố ý", "FAIL"],
    ["TC-64", "Đăng ký", "Khóa 'chờ thanh toán' bị ẩn khỏi Khóa của tôi", "1. API đăng ký khóa 2 với payment_status=pending\n2. GET /api/my-courses", "Đáng lẽ khóa hiện ra; THỰC TẾ bị ẩn (chỉ lọc completed)", "Lỗi cố ý", "FAIL"],

    # ----- Feature 4: Yêu thích -----
    ["TC-45", "Yêu thích", "Thêm vào yêu thích khi đã đăng nhập", "1. Đăng nhập\n2. Click nút tim khóa 1", "Toast thành công, tim chuyển active", "Bình thường", "PASS"],
    ["TC-46", "Yêu thích", "Thêm yêu thích khi chưa đăng nhập", "1. Click tim (chưa login)", "Hiện modal đăng nhập", "Bình thường", "PASS"],
    ["TC-47", "Yêu thích", "Xóa yêu thích từ trang chủ", "1. Thêm rồi click tim lần nữa", "Toast đã xóa, tim về trạng thái rỗng", "Bình thường", "PASS"],
    ["TC-48", "Yêu thích", "Trạng thái danh sách yêu thích trống", "1. Mở tab Yêu thích (chưa thích gì)", "Hiện 'Chưa có khóa học yêu thích nào'", "Bình thường", "PASS"],
    ["TC-49", "Yêu thích", "Hiển thị các khóa đã thích", "1. Thích khóa 1\n2. Mở tab Yêu thích", "Hiện 1 khóa 'Golang Cơ Bản'", "Bình thường", "PASS"],
    ["TC-50", "Yêu thích", "Mở chi tiết từ tab Yêu thích", "1. Click Chi tiết trong tab Yêu thích", "Mở modal chi tiết đúng khóa", "Bình thường", "PASS"],
    ["TC-51", "Yêu thích", "Đăng ký trực tiếp từ tab Yêu thích", "1. Thích khóa 2\n2. Đăng ký + thanh toán từ tab Yêu thích", "Badge 'Đã đăng ký'", "Bình thường", "PASS"],
    ["TC-52", "Yêu thích", "Bỏ thích ngay trong tab Yêu thích", "1. Click tim trong tab Yêu thích", "Thẻ biến mất, hiện trạng thái trống", "Bình thường", "PASS"],
    ["TC-53", "Yêu thích", "Đồng bộ trạng thái đăng ký giữa các tab", "1. Đăng ký từ tab Yêu thích\n2. Sang tab Khám phá", "Tab Khám phá cũng hiện 'Đã đăng ký'", "Bình thường", "PASS"],
    ["TC-54", "Yêu thích", "Class active/inactive của nút tim", "1. Kiểm tra class trước và sau khi click", "Đúng active khi đã thích, inactive khi chưa", "Bình thường", "PASS"],
    ["TC-55", "Yêu thích", "Đăng xuất ẩn tab Yêu thích", "1. Thích khóa\n2. Đăng xuất", "Tab Yêu thích biến mất khỏi navbar", "Bình thường", "PASS"],
    ["TC-56", "Yêu thích", "Cô lập wishlist giữa các tài khoản", "1. User A thích khóa 1\n2. User B thích khóa 2", "Mỗi user thấy đúng wishlist riêng của mình", "Bình thường", "PASS"],
    ["TC-57", "Yêu thích", "Badge đếm số lượng cập nhật động", "1. Thêm/bớt khóa yêu thích", "Badge cập nhật (0)→(1)→(2)→(1)", "Bình thường", "PASS"],
    ["TC-58", "Yêu thích", "Thêm lại khóa đã bỏ thích", "1. Thêm → Bỏ → Thêm lại", "Hoạt động trơn tru, thêm lại bình thường", "Bình thường", "PASS"],
    ["TC-59", "Yêu thích", "Spam click nút tim liên tục", "1. Click tim 2 lần liên tiếp", "Đáng lẽ cả 2 thành công; THỰC TẾ 1 request trả 500 (UNIQUE conflict)", "Lỗi cố ý", "FAIL"],
    ["TC-60", "Yêu thích", "Giữ wishlist khi chuyển tab", "1. Thích khóa 1\n2. Sang tab khác rồi quay lại", "Khóa vẫn còn trong tab Yêu thích", "Bình thường", "PASS"],
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Test Cases"

headers = ["Mã TC", "Nhóm tính năng", "Mục tiêu kiểm thử", "Các bước thực hiện", "Kết quả mong đợi", "Loại", "Kết quả"]

# Tiêu đề lớn
ws.merge_cells("A1:G1")
title_cell = ws["A1"]
title_cell.value = "DANH SÁCH TEST CASE - HỆ THỐNG KHÓA HỌC TRỰC TUYẾN TECHACADEMY"
title_cell.font = Font(bold=True, size=14, color="FFFFFF")
title_cell.fill = PatternFill("solid", fgColor="6D28D9")
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 28

total = len(data)
fail = sum(1 for r in data if r[6] == "FAIL")
ws.merge_cells("A2:G2")
sub = ws["A2"]
sub.value = f"Tổng số: {total} test case  |  PASS: {total - fail}  |  FAIL (lỗi cố ý): {fail}"
sub.font = Font(bold=True, size=11, color="111827")
sub.fill = PatternFill("solid", fgColor="EDE9FE")
sub.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 20

# Header
header_row = 3
for col, h in enumerate(headers, start=1):
    c = ws.cell(row=header_row, column=col, value=h)
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="8B5CF6")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.row_dimensions[header_row].height = 24

thin = Side(style="thin", color="D1D5DB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

fail_fill = PatternFill("solid", fgColor="FEE2E2")
pass_fill = PatternFill("solid", fgColor="ECFDF5")
bug_font = Font(bold=True, color="B91C1C")
pass_font = Font(bold=True, color="047857")

for i, row in enumerate(data):
    r = header_row + 1 + i
    is_fail = row[6] == "FAIL"
    for col, val in enumerate(row, start=1):
        c = ws.cell(row=r, column=col, value=val)
        c.border = border
        c.alignment = Alignment(vertical="top", wrap_text=True,
                                horizontal="center" if col in (1, 2, 6, 7) else "left")
        c.fill = fail_fill if is_fail else pass_fill
        if col == 7:
            c.font = bug_font if is_fail else pass_font

widths = [9, 14, 30, 38, 42, 12, 9]
for col, w in enumerate(widths, start=1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

ws.freeze_panes = "A4"

out = "Danh_sach_Test_Case_TECHACADEMY.xlsx"
wb.save(out)
print("Saved:", out, "| rows:", total, "| PASS:", total - fail, "| FAIL:", fail)
